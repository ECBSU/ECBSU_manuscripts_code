import openpyxl
import re
from openpyxl.styles import PatternFill

# Load workbook and sheet
wb = openpyxl.load_workbook("Orthogroups_extracted.xlsx")
ws = wb.active  # or use wb["SheetName"]

# Red fill style
red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

# Regex for gene code (e.g., ININ_####_)
gene_code_pattern = re.compile(r"\b[A-Z]{4}_")

# Loop through cells (excluding header)
for row in ws.iter_rows(min_row=2, min_col=2):  # adjust col range as needed
    for cell in row:
        matches = gene_code_pattern.findall(str(cell.value))
        if len(matches) > 1:
            cell.fill = red_fill

# Save updated file
wb.save("Orthogroups_extracted_dupliMarked.xlsx")
print("✅ Highlighted duplicated gene cells in Orthogroups_highlighted.xlsx")
