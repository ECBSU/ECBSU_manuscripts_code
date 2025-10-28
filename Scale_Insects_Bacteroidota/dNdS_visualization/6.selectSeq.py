import pandas as pd
from Bio import SeqIO
from Bio import pairwise2
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# ========== CONFIGURATION ==========
input_excel = "Orthogroups_extracted_dupliMarked_splitted_pseudoMarked.xlsx"
fasta_dir = "/Users/jinyeongchoi/Desktop/Projects/Bacteroidota/dN:dS_Symbiont_vs._free-living_bactrium_2nd/Extracted_AA_seq"  # ❗ Update this path
output_excel = "Orthogroups_extracted_dupliMarked_splitted_pseudoMarked_selected.xlsx"
log_file = "6.selectSeq.log"

symbiont_prefixes = [
    'ORUR', 'ICPU', 'ICSE', 'DRPI', 'DRCO', 'COCA', 'LLAX', 'ACLA',
    'ASSP', 'ININ', 'RASP', 'PUIS', 'AUIS', 'ASNE', 'COCH',
    'RIMA', 'RHAM', 'EUCO', 'RAIS', 'CRMU', 'CRGE'
]

# ========== READ INPUT ==========
df = pd.read_excel(input_excel)
selected_FLJO = {}
selected_FLPA = {}

with open(log_file, "w") as log:
    log.write("=== Selection Log: FLJO & FLPA (processed independently, single or duplicated genes) ===\n\n")
    log.write("Selection based on highest similarity (globalxx) between free-living genes and all symbiont sequences in the orthogroup fasta.\n")
    log.write("Similarity scores per comparison and average similarity scores per gene are reported below.\n\n")

    for _, row in df.iterrows():
        og = row["Orthogroup"]
        fasta_file = os.path.join(fasta_dir, f"{og}.faa")

        if not os.path.isfile(fasta_file):
            log.write(f"[{og}] ❌ FASTA not found: {fasta_file}\n\n")
            continue

        # Get FLJO and FLPA genes; keep only non-NaN strings
        FLJO_ids = [g for g in [row.get("FLJO_gene1"), row.get("FLJO_gene2")] if pd.notna(g)]
        FLPA_ids = [g for g in [row.get("FLPA_gene1"), row.get("FLPA_gene2")] if pd.notna(g)]

        # Skip if BOTH FLJO and FLPA genes are missing
        if len(FLJO_ids) == 0 and len(FLPA_ids) == 0:
            log.write(f"[{og}] Skipped: no FLJO or FLPA genes present\n\n")
            continue

        # Load sequences and find symbionts
        seqs = list(SeqIO.parse(fasta_file, "fasta"))
        symbiont_seqs = [s for s in seqs if any(s.id.startswith(prefix) for prefix in symbiont_prefixes)]
        if not symbiont_seqs:
            log.write(f"[{og}] ❌ No symbiont sequences found\n\n")
            continue

        def compare_all(gene_ids):
            results = {}  # gene_id -> list of (symbiont_id, score)
            for gid in gene_ids:
                gene_seq = next((s for s in seqs if s.id == gid), None)
                if not gene_seq:
                    results[gid] = []
                    continue
                scores = []
                for sym in symbiont_seqs:
                    score = pairwise2.align.globalxx(gene_seq.seq, sym.seq, one_alignment_only=True)[0].score
                    scores.append((sym.id, score))
                results[gid] = scores
            return results

        # Perform comparisons independently, whether single or duplicated
        FLJO_results = compare_all(FLJO_ids) if len(FLJO_ids) > 0 else {}
        FLPA_results = compare_all(FLPA_ids) if len(FLPA_ids) > 0 else {}

        def best_gene(results):
            all_scores = []
            for gid, sc_list in results.items():
                for sym_id, sc in sc_list:
                    all_scores.append((gid, sym_id, sc))
            if not all_scores:
                return None
            best = max(all_scores, key=lambda x: x[2])
            return best[0]

        selected_FLJO[og] = best_gene(FLJO_results) if FLJO_results else None
        selected_FLPA[og] = best_gene(FLPA_results) if FLPA_results else None

        # Logging FLJO results
        if FLJO_results:
            log.write(f"[{og}] FLJO comparisons:\n")
            for gid, sym_scores in FLJO_results.items():
                if sym_scores:
                    avg_score = sum(sc for _, sc in sym_scores) / len(sym_scores)
                    log.write(f"  {gid} average similarity score = {avg_score:.2f}\n")
                    for sym_id, sc in sym_scores:
                        log.write(f"    vs {sym_id}: Score -> Similarity score = {sc:.2f}\n")
                else:
                    log.write(f"  {gid} sequence not found or no symbiont comparison\n")
            log.write(f"  ✅ Selected FLJO gene: {selected_FLJO[og]}\n\n")
        else:
            log.write(f"[{og}] No FLJO genes to process\n\n")

        # Logging FLPA results
        if FLPA_results:
            log.write(f"[{og}] FLPA comparisons:\n")
            for gid, sym_scores in FLPA_results.items():
                if sym_scores:
                    avg_score = sum(sc for _, sc in sym_scores) / len(sym_scores)
                    log.write(f"  {gid} average similarity score = {avg_score:.2f}\n")
                    for sym_id, sc in sym_scores:
                        log.write(f"    vs {sym_id}: Score -> Similarity score = {sc:.2f}\n")
                else:
                    log.write(f"  {gid} sequence not found or no symbiont comparison\n")
            log.write(f"  ✅ Selected FLPA gene: {selected_FLPA[og]}\n\n")
        else:
            log.write(f"[{og}] No FLPA genes to process\n\n")

# ========== WRITE OUTPUT ==========
df.to_excel(output_excel, index=False)
wb = load_workbook(output_excel)
ws = wb.active
highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for row_num in range(2, ws.max_row + 1):
    og = ws.cell(row=row_num, column=1).value
    for col_num in range(2, ws.max_column + 1):
        cell = ws.cell(row=row_num, column=col_num)
        val = str(cell.value)
        if val in [selected_FLJO.get(og), selected_FLPA.get(og)]:
            cell.fill = highlight

wb.save(output_excel)

print("✅ Done!")
print(f"📘 Marked Excel: {output_excel}")
print(f"🪵 Log saved: {log_file}")
